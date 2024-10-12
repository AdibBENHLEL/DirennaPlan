import tkinter as tk
from tkinter import ttk, messagebox
#pour la date frame
from tkcalendar import DateEntry
#pour le live time
import time
#pour l'extract de pdf
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
#pour limportation de l'image 
from PIL import Image, ImageTk
import openpyxl
from openpyxl import load_workbook
import sys
import os
class TodoListApp:
    def __init__(self, root):
        
        
        self.root = root #This is the parent widget named root
        self.root.title("Direnna_Plan_App")
        
        # Set window size to 1/3 of screen width and full height
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        window_width = (screen_width // 3)-140
        window_height = screen_height-100
        self.root.geometry(f"{window_width}x{window_height}")
        
        self.root.config(bg="#2c3e50")
        
        if getattr(sys, 'frozen', False):
            # If the application is compiled as an .exe
            application_path = os.path.dirname(sys.executable)
        else:
            # If running in Python script mode
            application_path = os.path.dirname(__file__)

        # Path to the Excel file
        self.path_data = os.path.join(application_path, "data", "data_App.xlsx")
        # Try loading the workbook
        try:
            self.load_path_open_file = openpyxl.load_workbook(self.path_data)
            self.sheets = self.load_path_open_file["task_view"]
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Excel file: {e}")
            return
        # Path to the image
        self.image_path = os.path.join(application_path, "assets", "task.png")

        # Path to the app icon (when creating the window)
        self.root.iconbitmap(os.path.join(application_path, "assets", "icon.ico"))

        
        
        m_row = self.sheets.max_row
        self.tasks = []
        # Loop will print all values 
        for i in range(2, m_row+1):
                cell_obj_column1 = self.sheets.cell(row = i, column = 1)
                cell_obj_column2 = self.sheets.cell(row = i, column = 2)
                cell_obj_column4 = self.sheets.cell(row = i, column = 4)
                self.tasks.append((cell_obj_column1.value,cell_obj_column2.value,cell_obj_column4.value))#add tuple
                
        self.logo = tk.Label(root, text="DIRENNA_PLAN", font=("Helvetica", 24, "bold"), fg="#ecf0f1", bg="#2c3e50")
        self.logo.grid(row=0, column=0, padx=10, pady=10)

        # Load and display the image
        p=self.image_path
        image = Image.open(p)  # i;porter i;qge
        image = image.resize((100, 100), Image.LANCZOS)  # Resize the image
        
        self.photo = ImageTk.PhotoImage(image)
        self.image_label = tk.Label(root,image=self.photo, bg="#2c3e50")
        self.image_label.grid(row=0, column=1, padx=10, pady=10)

        self.label = tk.Label(root, text="Hello, Adib! Here are your tasks", font=("Helvetica", 16,"bold"), fg="#ecf0f1", bg="#2c3e50")
        self.label.grid(row=2, column=0, columnspan=2, pady=10, sticky="nsew")
        
        self.label_put_task = tk.Label(root, text="put your task", font=("Helvetica", 10,"bold"), fg="#ecf0f1", bg="#2c3e50")
        self.label_put_task.grid(row=3, column=1,padx=20, pady=10, sticky="nsew")
        
        self.task_entry = tk.Entry(root, width=20, font=("Helvetica", 16,"bold"))
        self.task_entry.grid(row=3, column=0, padx=20, pady=10, sticky="w")

        self.label_put_dead_line = tk.Label(root, text="put your deadline", font=("Helvetica", 10,"bold"), fg="#ecf0f1", bg="#2c3e50")
        self.label_put_dead_line.grid(row=4, column=1,padx=20, pady=10, sticky="nsew")
        
        self.cal = DateEntry(root, width=20, background="darkblue", foreground="white", borderwidth=3)
        self.cal.grid(row=4, column=0, padx=20, pady=10, sticky="nsew")
        
        add_button = tk.Button(
            root,
            width=20,
            text="Add Task",
            command=self.add_task,#la function exécuter si on click sur le button
            font=("Helvetica", 12,"bold"),
            bg="#27ae60",
            fg="#ecf0f1",
            activebackground="#2ecc71",
            activeforeground="#ecf0f1"
        )
        add_button.grid(row=5, column=0, columnspan=1, padx=20, pady=10, sticky="nsew")
        s = ttk.Style()
        s.configure('Treeview.Heading',font=('Arial', 10, 'bold'))
        
        self.checkbutton_value = tk.BooleanVar()
        c1 = tk.Checkbutton(root, text='done',variable=self.checkbutton_value,bg="#2c3e50", fg="#ecf0f1",font=("Helvetica", 12,"bold"),activebackground="#2c3e50",activeforeground="#ecf0f1",command=self.modify_prog_task)
        c1.grid(row=5, column=1, columnspan=1, padx=20, pady=10, sticky="nsew")
        
        self.task_listbox = ttk.Treeview(root, columns=("Task", "Deadline","Statue"), show="headings", height=8)# Treeview widget for displaying tabular data , option show="headlings" ensures that only the column headings are shown, and not the default tree structure
        self.task_listbox.heading("Task", text="Task")#put the text 
        self.task_listbox.heading("Deadline", text="Deadline")
        self.task_listbox.heading("Statue", text="Statue")

        self.task_listbox.column("Task", width=200)#put the width
        self.task_listbox.column("Deadline", width=160)
        self.task_listbox.column("Statue", width=90)

        self.task_listbox.grid(row=6, column=0, columnspan=2, padx=20, pady=20, sticky="nsew")#position of the treeview

        remove_button = tk.Button(
            root,
            width=15,
            text="Remove Task",
            command=self.remove_task,
            font=("Helvetica", 12,"bold"),
            bg="#e74c3c",
            fg="#ecf0f1",
            activebackground="#c0392b",
            activeforeground="#ecf0f1"
        )
        remove_button.grid(row=7, column=1,padx=20, pady=10, sticky="w")
        
        self.progression_input = tk.Entry(root, width=5, font=("Helvetica", 16,"bold"))
        self.progression_input.grid(row=7, column=0, padx=20, pady=10, sticky="e")
        
        modify_prog_button = tk.Button(
            root,
            width=15,
            text="modify_prog_Task",
            command=self.modify_prog_task,
            font=("Helvetica", 12,"bold"),
            bg="#f39c12",
            fg="#ecf0f1",
            activebackground="#e67e22",
            activeforeground="#ecf0f1"
        )
        modify_prog_button.grid(row=7, column=0, padx=20, pady=10, sticky="w")
        
        
        print_list_done_tasks_button = tk.Button(
            root,
            width=15,
            text="Print List Done Tasks",
            command=self.print_complete_task,
            font=("Helvetica", 12,"bold"),
            bg="#3498db",
            fg="#ecf0f1",
            activebackground="#2980b9",
            activeforeground="#ecf0f1"
        )
        print_list_done_tasks_button.grid(row=8, column=0, columnspan=1, padx=20, pady=10, sticky="nsew")

        self.task_listbox.bind("<Double-Button-1>", lambda event: self.show_details())#when i double click sur one row the fuction show_details will be excuted to show a message box

        # Add live date and time label
        self.time_label = tk.Label(root, text="", font=("Helvetica", 12,"bold"), bg="#2c3e50", fg="#ecf0f1")
        self.time_label.grid(row=1, column=0, padx=10, pady=10, sticky="e")
        self.update_time()#excuter la fonction pour faire update date

        self.la = tk.Label(root, text="All right reserved created by BEN HLEL adib ", font=("Helvetica", 10,"bold"), fg="#ecf0f1", bg="#2c3e50")
        self.la.grid(row=9, column=0, columnspan=2, pady=30, sticky="nsew")

        self.update_task_list()

    def update_time(self):
        self.current_time = time.strftime("%d-%m-%y %H:%M:%S")
        self.time_label.config(text=self.current_time)
        self.root.after(1000, self.update_time)

    def add_task(self):
        #file = open('data.xlsx', 'a')#append read mode 
        #file = csv.writer(file)
        task = self.task_entry.get()
        deadline = self.cal.get_date()
        #file.writerow([task, deadline,"12/1/2024","not done"])
        if task:
            self.tasks.append((task, deadline,"not done"))
            self.update_task_list()
            self.task_entry.delete(0, tk.END)
            self.sheets.append([task,deadline,self.current_time,"not done"])
            self.load_path_open_file.save(self.path_data)
        else:
            messagebox.showwarning("Warning", "Please enter a task.")

    def remove_task(self):
        selected_item = self.task_listbox.selection()
        if selected_item:
            res=messagebox.askquestion('warning', 'Do you really want to remove this task')
            if res == 'yes' :
                task_index = self.task_listbox.index(selected_item[0])
                index = self.task_listbox.index(selected_item)
            
                self.tasks.pop(task_index)
                self.sheets.delete_rows(index+2)

                self.update_task_list()
                self.load_path_open_file.save(self.path_data)
            
    def modify_prog_task(self):
        selected_item = self.task_listbox.selection()
        new_prog=self.progression_input.get()
        reponce_checkbutton=self.checkbutton_value.get()
        
        if ((new_prog or reponce_checkbutton )and selected_item) :
            task_index = self.task_listbox.index(selected_item[0])
            index = self.task_listbox.index(selected_item)
            
            task, deadline,Statue = self.tasks.pop(task_index)
            new_statue = new_prog if new_prog else ("done" if reponce_checkbutton else Statue)
            self.tasks.append((task, deadline,new_statue))
            self.progression_input.delete(0, tk.END)
            self.update_task_list()
            
            position_statue="D"+str(index+2)
            position_last_update="C"+str(index+2)
            #faire l'update de lexel por le champ statue et pour le shamp last_update
            self.sheets[position_statue]=new_statue
            self.sheets[position_last_update]=self.current_time
            self.load_path_open_file.save(self.path_data)
        else :
            messagebox.showwarning("Warning", "Please enter a new progression value for your task that you must select it.")

    def update_task_list(self):
        self.task_listbox.delete(*self.task_listbox.get_children())
        for task, deadline,Statue in self.tasks:
            self.task_listbox.insert("", tk.END, values=(task, deadline,Statue))
        '''
        for i in range(len(self.tasks)):
            if self.tasks[i][2] == "done":#champ statue
                couleur this row bg red with out tags
            else:
                couleur this row with bg bleu with out tags'''
    def print_complete_task(self):
        completed_tasks = [(task, deadline,Statue) for task, deadline,Statue in self.tasks if  Statue=="done" ]
        if not completed_tasks:
            messagebox.showinfo("Info", "No completed tasks to print.")
            return
        
        c = canvas.Canvas("Completed_Tasks.pdf", pagesize=letter)
        width, height = letter
        c.drawString(100, height - 100, "Completed Tasks")
        c.drawString(100, height - 120, "================")
        c.drawString(100, height - 50, "all rights reserved created by ben hlel adib")

        y = height - 140
        for task, deadline,Statue in completed_tasks:
            c.drawString(100, y, f"Task: {task}")
            c.drawString(100, y - 15, f"Date: {deadline}")
            c.drawString(100, y - 30, f"Status:{Statue}")
            c.drawString(100, y - 45, "-------------------------")
            y -= 60
        
        c.save()
        messagebox.showinfo("Info", "Completed tasks have been printed to Completed_Tasks.pdf.")
    def show_details(self):
        selected_item = self.task_listbox.selection()
        task_index = self.task_listbox.index(selected_item[0])
        index = self.task_listbox.index(selected_item)
        task, deadline,Statue = self.tasks.pop(task_index)
        messagebox.showinfo("Info", f"details about Task n° : {index} \n task : {task} \n deadline : {deadline} \n progression : {Statue}")

if __name__ == "__main__":
    root = tk.Tk()
    app = TodoListApp(root)
    root.mainloop()
